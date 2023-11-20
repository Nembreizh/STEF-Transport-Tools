#-------------------------------------------------------------------------------
# Name:        RepriseCA_XLS_SAS
# Purpose:
#
# Author:      Dreumont.b
#
# Created:     22/09/2023
# Copyright:   (c) Dreumont.b 2023
# Licence:     STEF TRANSPORT SAS
# Version:     1.1
# Modifications :
#   1.0     Programme initial
#   1.1     Ajout de la focntionnalité CSV
#-------------------------------------------------------------------------------

import tkinter as tk
from tkinter import filedialog, ttk
import csv
import openpyxl
import datetime

OUTPUT_COLUMNS = {
    "CE": "AD",
    "TYPE-ADR": "DE",
    "NTIERS": None,
    "NFL": None,
    "NTIERS-ext": None,
    "Nom": None,
    "Nom2": None,
    "Rue": None,
    "Rue2": None,
    "Ville": None,
    "CPostal": None,
    "Pays": None,
    "Code INSEE": None,
    "IDENT_CEE": None,
    "TELEPHONE": None,
    "FAX": None,
    "E_MAIL": None,
    "PDEP": None,
}

def browse_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("CSV Files", "*.csv")])
    if file_path.lower().endswith('.xlsx'):
        show_excel_preview(file_path)
    elif file_path.lower().endswith('.csv'):
        show_csv_preview(file_path)

def show_excel_preview(file_path):
    excel_file = openpyxl.load_workbook(file_path)
    sheet = excel_file.active
    header_row = next(sheet.iter_rows(values_only=True))
    column_names = [str(name) for name in header_row]
    show_mapping_window(column_names, sheet)

def show_csv_preview(file_path):
    global csv_content
    csv_content = []
    with open(file_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile, delimiter=';')
        header_row = next(reader)
        column_names = [str(name) for name in header_row]
        print("CSV Content:")
        for row in reader:
            csv_content.append(row)
        show_mapping_window(column_names, csv_content=csv_content)


def update_output_preview_csv():
    output_content_text.delete(1.0, tk.END)
    print("Updating CSV Preview:")
    for row in csv_content:
        print("Row:", row)
        csv_row = []
        for column_name, default_value in OUTPUT_COLUMNS.items():
            if column_mappings[column_name] is not None:
                csv_row.append(row[column_mappings[column_name]])
            elif default_value is not None:
                csv_row.append(default_value)
            else:
                if column_name in ["Rue", "CPostal", "Ville"]:
                    csv_row.append(".")
                else:
                    csv_row.append("")
        output_content_text.insert(tk.END, ";".join(map(str, csv_row)) + "\n")



def show_mapping_window(column_names, sheet=None, csv_content=None):
    root = tk.Tk()
    root.title("Convertisseur Excel/CSV vers fichier SAS AD")

    style = ttk.Style()
    style.configure("TButton", padding=(5, 5), width=20)

    column_mappings = {col: None for col in OUTPUT_COLUMNS.keys()}

    agence_var = tk.StringVar()
    donneur_ordre_var = tk.StringVar()
    def update_output_preview_csv():
        output_content_text.delete(1.0, tk.END)
        for row in csv_content:
            csv_row = []
            for column_name, default_value in OUTPUT_COLUMNS.items():
                if column_mappings[column_name] is not None and row[column_mappings[column_name]] != '':
                    csv_row.append(row[column_mappings[column_name]])
                elif default_value is not None:
                    csv_row.append(default_value)
                else:
                    if column_name in ["Rue", "CPostal", "Ville"]:
                        csv_row.append(".")
                    else:
                        csv_row.append("")
            output_content_text.insert(tk.END, ";".join(map(str, csv_row)) + "\n")


    def update_output_preview():
        output_content_text.delete(1.0, tk.END)
        if sheet:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                csv_row = []
                for column_name, default_value in OUTPUT_COLUMNS.items():
                    if column_mappings[column_name] is not None and row[column_mappings[column_name]] not in [None, 'None', '', 'NA', '#NA']:
                        csv_row.append(row[column_mappings[column_name]])
                    elif default_value is not None:
                        csv_row.append(default_value)
                    else:
                        if column_name in ["Rue", "CPostal", "Ville"]:
                            csv_row.append(".")
                        else:
                            csv_row.append("")
                output_content_text.insert(tk.END, ";".join(map(str, csv_row)) + "\n")
        elif csv_content:
            update_output_preview_csv()

    def convert():
        sas_file_path = filedialog.asksaveasfilename(defaultextension=".SAS", filetypes=[("Fichiers SAS", "*.SAS")])
        if sas_file_path:
            agence = agence_entry.get()[-3:]
            donneur_ordre = donneur_ordre_entry.get()
            current_date = datetime.datetime.now().strftime("%Y%m%d%H%M")

            with open(sas_file_path, 'w', newline='') as sasfile:
                sas_writer = csv.writer(sasfile, delimiter=';')
                sas_writer.writerow(['01', 'OT', '02', '', donneur_ordre, current_date, '', '', '', agence[-3:], donneur_ordre])
                sas_writer.writerow(['PO', 'C', agence[-3:], donneur_ordre, '', '', 'XX4', 'S', 'D', 'TM', '', '', '', '', '', '', '36011', '', '', 'O', '', '', '', donneur_ordre, '', donneur_ordre, current_date, current_date, '', '', '', '', '', '', donneur_ordre, 'FR22', '', 'O', '', '', donneur_ordre, 'FR', '', '', 'FR', '', '', '', '', '', '', '', '', '', '', '', '1', '000001', '', '0', '', '', '', '', '', '', '', '', '', '', '', 'PP', 'ME', '', '', '', '', '', '', '', '', '', '', '', '', 'CS'])
                if sheet:
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        sas_row = []
                        for column_name, default_value in OUTPUT_COLUMNS.items():
                            if column_mappings[column_name] is not None:
                                index = column_mappings[column_name]
                                value = str(row[index])[:35] if column_name in ["Nom", "Nom2", "Rue", "Rue2", "Ville"] else str(row[index])
                                sas_row.append(value)
                            elif default_value is not None:
                                sas_row.append(default_value)
                            else:
                                if column_name in ["Nom", "Rue", "CPostal", "Ville"]:
                                    sas_row.append(".")
                                else:
                                    sas_row.append("")
                        sas_writer.writerow(sas_row)
                    root.destroy()

    def update_mappings(event):
        for column_name, combobox in column_name_comboboxes.items():
            selected_column = combobox.get()
            if selected_column in column_names:
                column_mappings[column_name] = column_names.index(selected_column)
            else:
                column_mappings[column_name] = None
        update_output_preview()

    column_name_label = ttk.Label(root, text="Sélectionnez une colonne pour chaque champ:")
    column_name_label.grid(row=0, column=0, columnspan=2, padx=10, pady=10)

    column_name_comboboxes = {}

    for i, (column_name, default_value) in enumerate(OUTPUT_COLUMNS.items(), start=1):
        if column_name not in ['CE', 'TYPE-ADR']:
            label = ttk.Label(root, text=column_name)
            label.grid(row=i, column=0, padx=10, pady=5, sticky='w')

            column_name_combobox = ttk.Combobox(root, values=column_names)
            column_name_combobox.grid(row=i, column=1, padx=10, pady=5, sticky='w')

            if default_value is not None:
                column_name_combobox.set(default_value)

            column_name_comboboxes[column_name] = column_name_combobox
            column_name_combobox.bind("<<ComboboxSelected>>", update_mappings)

    agence_label = ttk.Label(root, text="Filiale (800XXX):")
    agence_label.grid(row=len(OUTPUT_COLUMNS) + 1, column=0, padx=10, pady=5, sticky='w')
    agence_entry = ttk.Entry(root, textvariable=agence_var)
    agence_entry.grid(row=len(OUTPUT_COLUMNS) + 1, column=1, padx=10, pady=5, sticky='w')

    donneur_ordre_label = ttk.Label(root, text="Donneur d'Ordre:")
    donneur_ordre_label.grid(row=len(OUTPUT_COLUMNS) + 2, column=0, padx=10, pady=5, sticky='w')
    donneur_ordre_entry = ttk.Entry(root, textvariable=donneur_ordre_var)
    donneur_ordre_entry.grid(row=len(OUTPUT_COLUMNS) + 2, column=1, padx=10, pady=5, sticky='w')

    convert_button = ttk.Button(root, text="Convertir en fichier SAS", command=convert)
    convert_button.grid(row=len(OUTPUT_COLUMNS) + 3, columnspan=2, padx=10, pady=10)

    output_preview_label = ttk.Label(root, text="Aperçu du fichier de sortie:")
    output_preview_label.grid(row=len(OUTPUT_COLUMNS) + 4, column=0, columnspan=2, padx=10, pady=10)

    global output_content_text
    output_content_text = tk.Text(root, height=10, width=80)
    output_content_text.grid(row=len(OUTPUT_COLUMNS) + 5, column=0, columnspan=2, padx=10, pady=10)

    update_output_preview()

    root.mainloop()

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Convertisseur Carnet d'adresse Client en fichier SAS avec ligne AD")

    browse_button = ttk.Button(root, text="Sélectionner un fichier Excel ou CSV", command=browse_file)
    browse_button.grid(row=0, column=0, padx=10, pady=10)

    root.mainloop()
